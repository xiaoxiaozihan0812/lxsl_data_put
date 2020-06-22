import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import shutil


def opener(row_num,column_num):
    '''
    1.更改下面漏扫记录单的路径以及名字
    :return:
    '''
    wb=load_workbook(filename=r'./漏扫记录单.xlsx')
    '''
    2.输入这个漏扫记录单的sheet1的名字
    '''
    sheet1=wb['5月中危联调']

    for i in range(1,int(row_num)+1):
        with open(r'.\loudongsaomiao/' + str(i) + '.txt', 'a+', encoding='utf-8') as f:
            f.write(sheet1[column_num + str(i)].value if sheet1[column_num + str(i)].value else f.write(''))

    for i in range(1,int(row_num)+1):
        with open(r'.\lousaonew/'+str(i)+'.txt','a+',encoding='utf-8') as f:
            f.read()

def searchip(row_num,row2_num):
    '''
    3.更改下面的路径和文件名，这个是要比对的那个自己整理好ip的表格。注意ip要在A列，也就是第一列
    '''
    wb = load_workbook(filename=r'.\ip地址.xlsx')
    '''
    4.更改下面sheet的名字
    '''
    sheet1 = wb['Sheet1']
    for yea in range(1,int(row_num)+1):
        print('正在检索第',str(yea),'格数据')
        f=open('.\loudongsaomiao/'+str(yea)+'.txt','r',encoding='utf-8')
        ff=open('.\lousaonew/'+str(yea)+'.txt','a+',encoding='utf-8')
        p=f.readlines()
        # print(p)
        p = [x.strip() for x in p if x.strip() != '']
        # print(p)
        for i in p[:]:
            # print(i)
            for j in range(1,int(row2_num)+1):
                # print(sheet1['A'+str(j)].value)
                if sheet1['A'+str(j)].value == i:
                    print('找到应该删除的ip：'+sheet1['A'+str(j)].value)
                    ff.write('#')
                    try:
                        p.remove(i)
                    except Exception as e:
                        print(e)
        for x in p:
            ff.write(x+'\n')

        ff.close()
        f.close()


def writetxt(row_num,column_num):
    '''
    5.更改下面漏扫记录单的路径以及名字
    '''
    wb = load_workbook(filename=r'./漏扫记录单.xlsx')
    '''
    6.输入这个漏扫记录单的sheet1的名字
    '''
    sheet1 = wb['5月中危联调']
    for i in range(1,int(row_num)+1):
        f=open('.\lousaonew/'+str(i)+'.txt','r',encoding='utf-8')
        c=f.read()
        # print(c)
        sheet1[column_num+str(i)]=c
        print('正在写入第:',str(i))
        '''
         7.更改下面漏扫记录单的路径以及名字,和上面一样

         '''
        wb.save(r'.\漏扫记录单.xlsx')
        f.close()
def Delete_File_Dir(dirName):
    shutil.rmtree(dirName)

if __name__ == '__main__':
    os.makedirs('.\loudongsaomiao')
    os.makedirs('.\lousaonew')
    '''
    8.输入漏漏扫记录单想要获取多少行数据
      '''
    row_num = input('输入漏扫整改记录单表格里想要获取ip那列多少行的数据:')
    '''
    9：输入漏扫记录单的ip在哪列，比如在C列就填写C
    '''
    column_num = input('输入漏扫整改记录单的ip列数，在哪列就填写哪列:').upper()
    '''

    10:输入整理好的要比对的Ip的表格一共的行数
    '''
    row2_num = input('请输入整理好的准备去比对的那个自己整理的Ip表格一共的行数：')
    opener(row_num, column_num)
    searchip(row_num, row2_num)
    writetxt(row_num, column_num)

    Delete_File_Dir(r'.\loudongsaomiao')
    Delete_File_Dir(r'.\lousaonew')

#236 G 5月中危运行 19061
#135 G  5月中危联调 19061
